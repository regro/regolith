"""Builder for Recent Collaborators."""
import datetime as dt
import os
import sys
from copy import copy

import openpyxl
from dateutil.relativedelta import relativedelta
from nameparser import HumanName

from regolith.builders.basebuilder import BuilderBase
from regolith.dates import month_to_int
from regolith.sorters import position_key
from regolith.tools import all_docs_from_collection, filter_publications, \
    fuzzy_retrieval, is_since

NUM_MONTHS = 48


def mdy_date(month, day, year):
    """Make a date object."""
    if isinstance(month, str):
        month = month_to_int(month)
    return dt.date(year, month, day)


def mdy(month, day, year):
    """Format the date to a string mm/dd/yy."""
    return "{}/{}/{}".format(
        str(month_to_int(month)).zfill(2), str(day).zfill(2), str(year)[-2:]
    )


def get_advisors_name_inst(advisee, rc):
    """Get the advisee's advisor. Yield (last name, first name, institution name)."""
    my_eme = advisee.get("employment", []) + advisee.get("education", [])
    relevant_emes = [i for i in my_eme if "advisor" in i]
    phd_advisors = [
        (i.get("advisor"), "phd")
        for i in relevant_emes
        if 'phd' or "dphil" in i.get("degree", "").lower()
    ]
    pdoc_advisors = [
        (i.get("advisor"), "postdoc")
        for i in relevant_emes if "organization" in i
    ]
    advisors = phd_advisors + pdoc_advisors
    for advisor in advisors:
        adv = fuzzy_retrieval(
            all_docs_from_collection(rc.client, "contacts"),
            ['aka', 'name', '_id'], advisor[0],
            case_sensitive=False
        )
        if adv:
            advsior_name = HumanName(adv.get("name"))
            inst = fuzzy_retrieval(
                all_docs_from_collection(rc.client, "institutions"),
                ['aka', 'name', '_id'], adv.get("institution"),
                case_sensitive=False
            )
            if inst:
                yield advsior_name.last, advsior_name.first, inst.get("name", "")
            else:
                print("WARNING: {} not in institutions".format(adv.get("institution")))
                yield advsior_name.last, advsior_name.first, adv.get("institution")


def get_advisees_name_inst(coll, advisor, rc):
    """Get advisor's advisees. Yield (last name, first name, institutions)"""
    advisor_names = advisor.get('aka', []) + [advisor.get('name'), advisor.get('_id')]
    for person in coll:
        edus = person.get("education", [])
        for edu in edus:
            if 'advisor' in edu and edu['advisor'] in advisor_names:
                person_name = HumanName(person.get("name"))
                inst_name = edu.get("institution")
                inst = fuzzy_retrieval(
                    all_docs_from_collection(rc.client, "institutions"),
                    ['aka', 'name', '_id'], inst_name,
                    case_sensitive=False)
                if inst is None:
                    print("WARNING: {} not in institutions".format(
                        inst_name))
                    yield person_name.last, person_name.first, inst_name
                else:
                    yield person_name.last, person_name.first, inst.get('name', "")
                break


def filter_since_date(pubs, since_date):
    """Filter the publications after the since_date."""
    for pub in pubs:
        if is_since(pub.get("year"), since_date.year,
                    pub.get("month", 1), since_date.month):
            if not pub.get("month"):
                print("WARNING: {} is missing month".format(
                    pub["_id"]))
            if pub.get("month") == "tbd".casefold():
                print("WARNING: month in {} is tbd".format(
                    pub["_id"]))
            yield pub


def get_since_date(rc):
    """Get the since_date from command."""
    if isinstance(rc, str):
        since_date = dt.datetime.strptime(rc, '%Y-%m-%d').date() - relativedelta(months=NUM_MONTHS)
    else:
        since_date = dt.date.today() - relativedelta(months=NUM_MONTHS)
    return since_date


def get_coauthors_from_pubs(pubs):
    """Get co-authors' names from the publication."""
    my_collabs = []
    for pub in pubs:
        my_collabs.extend(
            [
                collabs for collabs in
                (names for names in pub.get('author', []))
            ]
        )
    return my_collabs


def get_recent_org(person_info):
    """Get the person's most recent organization."""
    if "employment" in person_info:
        employment = person_info.get("employment", []) + person_info.get("education", [])
        if len(employment) == 0:
            return ""
        # sort by end_year
        employment = sorted(
            employment,
            key=lambda d: d.get("end_year", float('inf')),
            reverse=True)
        organization = employment[0].get('organization', '')
    else:
        organization = ""
    return organization


def query_people_and_institutions(rc, names):
    """Get the people and institutions names."""
    people, institutions = [], []
    for person_name in names:
        person_found = fuzzy_retrieval(all_docs_from_collection(
            rc.client, "people"),
            ["name", "aka", "_id"],
            person_name, case_sensitive=False)
        if not person_found:
            person_found = fuzzy_retrieval(all_docs_from_collection(
                rc.client, "contacts"),
                ["name", "aka", "_id"], person_name, case_sensitive=False)
            if not person_found:
                print(
                    "WARNING: {} not found in contacts or people. Check aka".format(
                        person_name))
            else:
                people.append(person_found['name'])
                inst = fuzzy_retrieval(all_docs_from_collection(
                    rc.client, "institutions"),
                    ["name", "aka", "_id"],
                    person_found["institution"], case_sensitive=False)
                if inst:
                    institutions.append(inst["name"])
                else:
                    institutions.append(person_found.get("institution", "missing"))
                    print("WARNING: {} missing from institutions".format(
                        person_found["institution"]))
        else:
            people.append(person_found['name'])
            pinst = get_recent_org(person_found)
            inst = fuzzy_retrieval(all_docs_from_collection(
                rc.client, "institutions"), ["name", "aka", "_id"],
                pinst, case_sensitive=False)
            if inst:
                institutions.append(inst["name"])
            else:
                institutions.append(pinst)
                print(
                    "WARNING: {} missing from institutions".format(
                        pinst))
    return people, institutions


def get_inst_name(person, rc):
    """Get the name of instituion of the person's lastest employment."""
    if 'employment' in person:
        org = get_recent_org(person)
        person_inst_abbr = org
    elif 'institution' in person:
        person_inst_abbr = person.get('institution')
    else:
        person_inst_abbr = ''
    person_inst = fuzzy_retrieval(all_docs_from_collection(
        rc.client, "institutions"), ["name", "aka", "_id"],
        person_inst_abbr, case_sensitive=False)
    if person_inst is not None:
        person_inst_name = person_inst.get("name")
    else:
        person_inst_name = person_inst_abbr
        print(f"WARNING: {person_inst_abbr} is not found in institutions.")
    return person_inst_name


def get_person_pubs(coll, person):
    """Get the publications from one person."""
    my_names = frozenset(person.get("aka", []) + [person["name"]])
    pubs = filter_publications(
        coll,
        my_names,
        reverse=True,
        bold=False
    )
    return pubs


def make_person_3tups(person, rc):
    if 'name' not in person:
        print("Warning")
    name = HumanName(person['name'])
    inst = get_inst_name(person, rc)
    return [(name.last, name.first, inst)]


def format_last_first_instutition_names(rc, ppl_names, excluded_inst_name=None):
    """Get the last name, first name and institution name."""
    ppl = []
    for ppl_tup in ppl_names:
        inst = fuzzy_retrieval(
            all_docs_from_collection(rc.client, "institutions"),
            ['aka', 'name', '_id'], ppl_tup[1],
            case_sensitive=False)
        if inst:
            inst_name = inst.get("name", "")
        else:
            inst_name = ppl_tup[1]
        # remove all people who are in the institution of the person
        if inst_name != excluded_inst_name:
            name = HumanName(ppl_tup[0])
            yield name.last, name.first, ppl_tup[1]
    return ppl


def format_people_name(ppl_names):
    """Format people name to be first name, last name."""
    ppl = set()
    # reformatting the name in last name, first name
    for idx in range(len(ppl_names)):
        names = ppl_names[idx][0].split()
        last_name = names[-1]
        first_name = ' '.join(names[:-1])
        name_reformatted = ', '.join([last_name, first_name])
        ppl.add((name_reformatted, ppl_names[idx][1]))
    return list(ppl)


def format_to_nsf(tups, type_str):
    """Format the 3 tups to 2 tups. ('type_str', 'last, first', 'inst', ...)."""
    return [
        (type_str, '{}, {}'.format(tup[0], tup[1])) + tup[2:] for tup in tups
    ]


def apply_cell_style(*cells, style):
    """Apply the format to cells."""
    for cell in cells:
        cell.font = style["font"]
        cell.border = style["border"]
        cell.fill = style["fill"]
        cell.alignment = style["alignment"]
    return


def copy_cell_style(style_ref_cell):
    """Copy the cell format to a dictionary."""
    template_cell_style = {
        "font": copy(style_ref_cell.font),
        "border": copy(style_ref_cell.border),
        "fill": copy(style_ref_cell.fill),
        "alignment": copy(style_ref_cell.alignment)
    }
    return template_cell_style


def is_merged(cell):
    """Whether the cell is merged."""
    return True if type(cell).__name__ == 'MergedCell' else False


def find_merged_cell(cells):
    """Find the index of all merged cells."""
    i, j = 0, 0
    lst = []
    while i < len(cells):
        if is_merged(cells[i]):
            while j < len(cells):
                if is_merged(cells[j]):
                    j += 1
                else:
                    i = j
                    break
            lst.append((i, j - 1))
        else:
            i += 1
    return lst


def unmerge(ws, cells):
    """Unmerge the cells."""
    lst = find_merged_cell(cells)
    for start, end in lst:
        ws.unmerge_cells("{}:{}").format(cells[start].coordinate, cells[end].coordinate)
    return


def get_person(person_id, rc):
    """Get the person's name."""
    person_found = fuzzy_retrieval(
        all_docs_from_collection(rc.client, "people"),
        ["name", "aka", "_id"],
        person_id,
        case_sensitive=False
    )
    if person_found:
        return person_found
    person_found = fuzzy_retrieval(
        all_docs_from_collection(rc.client, "contacts"),
        ["name", "aka", "_id"],
        person_id,
        case_sensitive=False
    )
    if person_found:
        return person_found
    print("WARNING: {} missing from people and contacts. Check aka.".format(person_id))
    return None


def find_coeditors(person, rc):
    """Get the coeditors info of the person. Return (last, first, inst, journal)."""
    emps = person.get('employment')
    if emps is None:
        return set()

    def coeditor_id_journals(_emps):
        for emp in _emps:
            if emp.get('position') == 'editor':
                _journal = emp.get('department', '')
                coeditor_ids = emp.get('coworkers', [])
                for _coeditor_id in coeditor_ids:
                    yield _coeditor_id, _journal

    coeditor_inst_journals = set()
    for coeditor_id, journal in coeditor_id_journals(emps):
        coeditor = get_person(coeditor_id, rc)
        coeditor_name = HumanName(coeditor['name'])
        inst_name = get_inst_name(coeditor, rc)
        coeditor_inst_journals.add((coeditor_name.last, coeditor_name.first, inst_name, journal))
    return coeditor_inst_journals


class RecentCollaboratorsBuilder(BuilderBase):
    """Build recent collaborators from database entries"""
    btype = "recent-collabs"
    needed_dbs = ['citations', 'people', 'contacts', 'institutions']

    def __init__(self, rc):
        """Initiate the class instance."""
        super().__init__(rc)
        self.template = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "templates", "coa_template_nsf.xlsx"
        )
        self.template2 = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "templates", "coa_template_doe.xlsx"
        )
        self.cmds = ["excel"]

    def construct_global_ctx(self):
        """Construct the global ctx including database and methods."""
        super().construct_global_ctx()
        gtx = self.gtx
        rc = self.rc
        gtx["people"] = sorted(
            all_docs_from_collection(rc.client, "people"),
            key=position_key,
            reverse=True,
        )
        gtx["contacts"] = sorted(
            all_docs_from_collection(rc.client, "contacts"),
            key=position_key,
            reverse=True,
        )
        gtx["institutions"] = all_docs_from_collection(rc.client,
                                                       "institutions")
        gtx["citations"] = all_docs_from_collection(rc.client, "citations")
        gtx["all_docs_from_collection"] = all_docs_from_collection

    def query_ppl(self, target, **filters):
        """Query the data base for the target's collaborators' information."""
        rc = self.rc
        gtx = self.gtx
        person = fuzzy_retrieval(all_docs_from_collection(rc.client, "people"),
                                 ['aka', 'name', '_id'], target,
                                 case_sensitive=False)
        if not person:
            raise RuntimeError("Person {} not found in people.".format(target))
        pubs = get_person_pubs(gtx["citations"], person)
        if 'since_date' in filters:
            since_date = filters.get('since_date')
            pubs = filter_since_date(pubs, since_date)
        my_collabs = get_coauthors_from_pubs(pubs)
        people, institutions = query_people_and_institutions(rc, my_collabs)
        ppl_names = set(zip(people, institutions))
        collab_3tups = set(format_last_first_instutition_names(rc, ppl_names))
        advisors_3tups = set(get_advisors_name_inst(person, rc))
        advisees_3tups = set(get_advisees_name_inst(gtx["people"], person, rc))
        ppl_3tups = sorted(list(collab_3tups | advisors_3tups | advisees_3tups))
        person_3tups = make_person_3tups(person, rc)
        coeditors_info = find_coeditors(person, rc)
        ppl_tab1 = format_to_nsf(person_3tups, '')
        ppl_tab3 = format_to_nsf(advisors_3tups, 'G:') + format_to_nsf(advisees_3tups, 'T:')
        ppl_tab4 = format_to_nsf(collab_3tups, 'A:')
        ppl_tab5 = format_to_nsf(coeditors_info, 'E:')
        results = {
            'person_info': person,
            'ppl_tab1': ppl_tab1,
            'ppl_tab3': ppl_tab3,
            'ppl_tab4': ppl_tab4,
            'ppl_tab5': ppl_tab5,
            'ppl_3tups': ppl_3tups
        }
        return results

    @staticmethod
    def fill_in_tab(ws, ppl_tups, start_row, template_cell_style=None, cols='ABCDE'):
        """Add the information in person, institution pairs into the table 4 in nsf table."""
        more_rows = len(ppl_tups) - 1
        if more_rows > 0:
            ws.insert_rows(start_row, amount=more_rows)
        for row, tup in enumerate(ppl_tups, start=start_row):
            cells = [ws['{}{}'.format(col, row)] for col in cols]
            if template_cell_style is not None:
                apply_cell_style(*cells, style=template_cell_style)
            for ind, value in enumerate(tup):
                cells[ind].value = value
        return

    def render_template1(self, person_info, ppl_tab1, ppl_tab3, ppl_tab4, ppl_tab5, **kwargs):
        """Render the nsf template."""
        template = self.template
        wb = openpyxl.load_workbook(template)
        ws = wb.worksheets[0]
        style = copy_cell_style(ws['A17'])
        self.fill_in_tab(
            ws, ppl_tab5, start_row=44, template_cell_style=style
        )
        self.fill_in_tab(
            ws, ppl_tab4, start_row=37, template_cell_style=style
        )
        self.fill_in_tab(
            ws, ppl_tab3, start_row=30, template_cell_style=style
        )
        self.fill_in_tab(
            ws, ppl_tab1, start_row=17, template_cell_style=style
        )
        wb.save(os.path.join(self.bldir, "{}_nsf.xlsx".format(person_info["_id"])))
        return locals()

    def render_template2(self, person_info, ppl_3tups, **kwargs):
        """Render the doe template."""
        template2 = self.template2
        ppl_3tups = ppl_3tups
        num_rows = len(ppl_3tups)
        wb = openpyxl.load_workbook(template2)
        ws = wb.worksheets[0]
        for row in range(num_rows):
            ws["A{}".format(row + 8)].value = ppl_3tups[row][0]
            ws["B{}".format(row + 8)].value = ppl_3tups[row][1]
            ws["C{}".format((row + 8))].value = ppl_3tups[row][2]
        wb.save(os.path.join(self.bldir, "{}_doe.xlsx".format(person_info["_id"])))
        return locals()

    def excel(self):
        """Query data base and build nsf and doe excels."""
        rc = self.rc
        if not rc.people:
            sys.exit("please rerun specifying --people PERSON")
        if isinstance(rc.people, str):
            rc.people = [rc.people]
        since_date = get_since_date(rc)
        target = rc.people[0]
        query_results = self.query_ppl(target, since_date=since_date)
        self.render_template1(**query_results)
        self.render_template2(**query_results)
