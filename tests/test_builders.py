import os

# from xonsh.lib import subprocess
import subprocess
import sys
from pathlib import Path

import habanero
import openpyxl
import pytest

from regolith.broker import load_db
from regolith.main import main

builder_map = [
    "cv",
    "review-prop",
    "annual-activity",
    "beamplan",
    "current-pending",
    "figure",
    "formalletter",
    #    "html",
    "internalhtml",
    "preslist",
    "publist",
    "releaselist",
    "recent-collabs",
    "grant-report",
    "resume",
    "review-man",
    # reading-lists    need tests for this
    "reimb",
]
db_srcs = ["mongo", "fs"]

xls_check = ("B17", "B20", "B36")
recent_collabs_xlsx_check = ["A51", "B51", "C51"]


def is_same(text0: str, text1: str, ignored: list):
    """Compare the content of two text.

    If there are different words in text0 and text1 and the word in
    text0 does not contain ignored substring, return False. Else return
    True.
    """

    def should_ignore(word):
        for w in ignored:
            if w in word:
                return True
        return False

    words0, words1 = text0.split(), text1.split()
    if len(words0) != len(words1):
        return False
    for word0, word1 in zip(words0, words1):
        if not should_ignore(word0) and word0 != word1:
            return False
    return True


def prep_figure():
    # Make latex file with some jinja2 in it
    text = r"""
    \include{ {{-get_file_path(db['groups']['ergs'], 'hello')-}}}"""
    with open("figure.tex", "w") as f:
        f.write(text)
    # make file to be loaded
    os.makedirs("fig", exist_ok=True)
    with open("fig/hello.txt", "w") as f:
        f.write("hello world")
    # load the db and register the file
    db = load_db()
    print(db.get_file_path(db["groups"]["ergs"], "hello"))
    if not db.get_file_path(db["groups"]["ergs"], "hello"):
        db.add_file(db["groups"]["ergs"], "hello", "fig/hello.txt")


@pytest.mark.skipif(sys.platform == "win32", reason="does not run on windows")
@pytest.mark.parametrize("bm", builder_map)
@pytest.mark.parametrize("db_src", db_srcs)
def test_builder(bm, db_src, make_db, make_mongodb, monkeypatch):
    # FIXME: Somehow the mongo backend failed to build figure
    # FIXME: now fs is failing to build figure

    # if db_src == "mongo" and bm == "figure":
    if bm == "figure":
        return
    if db_src == "fs":
        repo = make_db
    elif db_src == "mongo":
        if make_mongodb is False:
            pytest.skip("Mongoclient failed to start")
        else:
            repo = make_mongodb
    else:
        raise ValueError("Unknown database source: {}".format(db_src))
    os.chdir(repo)
    if bm == "figure":
        prep_figure()
    if bm == "internalhtml":
        # for some reason the mocking of the crossref call doesn't work when the
        # test is run using subprocess, so skip in this case.
        # the functionality is fully tested in test_builder_python
        pytest.skip("mocking of Crossref not working with subprocess")
    if bm == "html":
        os.makedirs("templates/static", exist_ok=True)
    if bm == "reimb" or bm == "recent-collabs":
        subprocess.run(["regolith", "build", bm, "--no-pdf", "--people", "scopatz"], check=True, cwd=repo)
    elif bm == "annual-activity":
        subprocess.run(
            ["regolith", "build", bm, "--no-pdf", "--people", "sbillinge", "--from", "2017-04-01"],
            check=True,
            cwd=repo,
        )
    elif bm == "grant-report":
        main(["build", bm, "--no-pdf", "--grant", "SymPy-1.1", "--from", "2017-04-01", "--to", "2018-03-31"])
    elif bm == "publist":
        subprocess.run(["regolith", "build", bm, "--no-pdf"], check=True, cwd=repo)
        subprocess.run(
            [
                "regolith",
                "build",
                bm,
                "--no-pdf",
                "--kwargs",
                "facility:nslsii",
                "--from",
                "2016-01-01",
                "--to",
                "2018-12-01",
                "--people",
                "scopatz",
            ],
            check=True,
            cwd=repo,
        )
    else:
        subprocess.run(["regolith", "build", bm, "--no-pdf"], check=True, cwd=repo)
    repo_dir = Path(repo)
    build_dir = repo_dir / "_build" / bm
    os.chdir(build_dir)
    # find all the files in the test outputs
    test_outputs_dir = Path(__file__).parent / "outputs" / bm
    all_test_items = test_outputs_dir.rglob("*")
    test_files = [file for file in all_test_items if file.is_file()]
    # find all the files output by the function
    all_outputs = build_dir.rglob("*")
    output_files = [file for file in all_outputs if file.is_file()]
    output_filenames = [file.name for file in output_files]
    # if there is a test output file that wasn't produced by the code, fail test
    for testfile in test_files:
        if testfile.name not in output_filenames:
            print(
                f"Expected test file {testfile} not generated by the builder." f"Files generated: {output_files}"
            )
            assert False
    # for all other test output 'expected' files check they are the same as the program output
    for file in test_files:
        actual_file = build_dir / file.name
        # if they are tex or html, disregard certain formatting things
        html_tex_bool = False
        if file.name.endswith(".html") or file.name.endswith(".tex"):
            html_tex_bool = True
        if bm == "reimb":
            actual = openpyxl.load_workbook(actual_file)["T&B"]
            actual = [str(actual[b]) for b in xls_check]
        elif bm == "recent-collabs":
            if "nsf" in actual_file.name:
                sheet = "NSF COA Template"
            else:
                sheet = "Collaborators"
            actual = openpyxl.load_workbook(actual_file)[sheet]
            actual = [str(actual[cell]) for cell in recent_collabs_xlsx_check]
        elif html_tex_bool:
            with open(actual_file, "r") as f:
                actual = [line.strip() for line in f]
            actual = [string for string in actual if ".." not in string]
            actual = [string for string in actual if "Temp" not in string]
            actual = [string for string in actual if "/tmp/" not in string]
            actual = [string for string in actual if len(string) != 0]
        else:
            with open(actual_file, "r") as f:
                actual = f.read()
        if bm == "reimb":
            expected = openpyxl.load_workbook(file)["T&B"]
            expected = [str(expected[b]) for b in xls_check]
        elif bm == "recent-collabs":
            if "nsf" in file.name:
                sheet = "NSF COA Template"
            else:
                sheet = "Collaborators"
            expected = openpyxl.load_workbook(file)[sheet]
            expected = [str(expected[cell]) for cell in recent_collabs_xlsx_check]
        elif html_tex_bool:
            with open(file, "r") as f:
                expected = [line.strip() for line in f]
            expected = [string for string in expected if ".." not in string]
            expected = [string for string in expected if "Temp" not in string]
            expected = [string for string in expected if "/tmp/" not in string]
            expected = [string for string in expected if len(string) != 0]
        else:
            with open(file, "r") as f:
                expected = f.read()

        # Skip because of a date time in
        if file != "rss.xml":
            assert expected == actual


@pytest.mark.parametrize("db_src", db_srcs)
@pytest.mark.parametrize("bm", builder_map)
def test_builder_python(bm, db_src, make_db, make_mongodb, monkeypatch):
    # FIXME: Somehow the mongo backend failed to build figure
    # FIXME: now fs is failing to build figure
    # if db_src == "mongo" and bm == "figure":
    if bm == "figure":
        return
    if db_src == "fs":
        repo = make_db
    elif db_src == "mongo":
        if make_mongodb is False:
            pytest.skip("Mongoclient failed to start")
        else:
            repo = make_mongodb
    os.chdir(repo)
    if bm == "figure":
        prep_figure()
    if bm == "internalhtml":

        def mockreturn(*args, **kwargs):
            mock_article = {
                "message": {
                    "author": [{"given": "SJL", "family": "B"}],
                    "short-container-title": ["J Club Paper"],
                    "volume": 10,
                    "title": ["title"],
                    "issued": {"date-parts": [[1971]]},
                }
            }
            return mock_article

        monkeypatch.setattr(habanero.Crossref, "works", mockreturn)
    if bm == "html":
        os.makedirs("templates/static", exist_ok=True)
    if bm == "reimb" or bm == "recent-collabs":
        main(["build", bm, "--no-pdf", "--people", "scopatz"])
    elif bm == "annual-activity":
        main(["build", bm, "--no-pdf", "--people", "sbillinge", "--from", "2017-04-01"])
    elif bm == "grant-report":
        main(["build", bm, "--no-pdf", "--grant", "SymPy-1.1", "--from", "2017-04-01", "--to", "2018-03-31"])
    elif bm == "publist":
        main(["build", bm, "--no-pdf"])
        main(
            [
                "build",
                bm,
                "--no-pdf",
                "--kwargs",
                "facility:nslsii",
                "--from",
                "2016-01-01",
                "--to",
                "2018-12-01",
                "--people",
                "scopatz",
            ]
        )
    else:
        main(["build", bm, "--no-pdf"])
    repo_dir = Path(repo)
    build_dir = repo_dir / "_build" / bm
    os.chdir(build_dir)
    # find all the files in the test outputs
    test_outputs_dir = Path(__file__).parent / "outputs" / bm
    all_test_items = test_outputs_dir.rglob("*")
    test_files = [file for file in all_test_items if file.is_file()]
    # find all the files output by the function
    all_outputs = build_dir.rglob("*")
    output_files = [file for file in all_outputs if file.is_file()]
    output_filenames = [file.name for file in output_files]
    # if there is a test output file that wasn't produced by the code, fail test
    for testfile in test_files:
        if testfile.name not in output_filenames:
            print(
                f"Expected test file {testfile} not generated by the builder." f"Files generated: {output_files}"
            )
            assert False
    # for all other test output 'expected' files check they are the same as the program output
    for file in test_files:
        actual_file = build_dir / file.name
        # if they are tex or html, disregard certain formatting things
        html_tex_bool = False
        if file.name.endswith(".html") or file.name.endswith(".tex"):
            html_tex_bool = True
        if bm == "reimb":
            actual = openpyxl.load_workbook(actual_file)["T&B"]
            actual = [str(actual[b]) for b in xls_check]
        elif bm == "recent-collabs":
            if "nsf" in actual_file.name:
                sheet = "NSF COA Template"
            else:
                sheet = "Collaborators"
            actual = openpyxl.load_workbook(actual_file)[sheet]
            actual = [str(actual[cell]) for cell in recent_collabs_xlsx_check]
        elif html_tex_bool:
            with open(actual_file, "r") as f:
                actual = [line.strip() for line in f]
            actual = [string for string in actual if ".." not in string]
            actual = [string for string in actual if "Temp" not in string]
            actual = [string for string in actual if "/tmp/" not in string]
            actual = [string for string in actual if len(string) != 0]
        else:
            with open(actual_file, "r") as f:
                actual = f.read()

        if bm == "reimb":
            expected = openpyxl.load_workbook(file)["T&B"]
            expected = [str(expected[b]) for b in xls_check]
        elif bm == "recent-collabs":
            if "nsf" in file.name:
                sheet = "NSF COA Template"
            else:
                sheet = "Collaborators"
            expected = openpyxl.load_workbook(file)[sheet]
            expected = [str(expected[cell]) for cell in recent_collabs_xlsx_check]
        elif html_tex_bool:
            with open(file, "r") as f:
                expected = [line.strip() for line in f]
            expected = [string for string in expected if ".." not in string]
            expected = [string for string in expected if "Temp" not in string]
            expected = [string for string in expected if "/tmp/" not in string]
            expected = [string for string in expected if len(string) != 0]
        else:
            with open(file, "r") as f:
                expected = f.read()

        # Skip because of a date time in
        if file != "rss.xml":
            assert expected == actual
