"""Builder for Resumes."""

import os

import openpyxl

from regolith.builders.basebuilder import BuilderBase
from regolith.dates import get_dates
from regolith.sorters import position_key
from regolith.tools import all_docs_from_collection, fuzzy_retrieval, month_and_year


class ReimbursementBuilder(BuilderBase):
    """Build reimbursement from database entries."""

    btype = "reimb"
    needed_colls = ["expenses", "people", "grants"]

    def __init__(self, rc):
        super().__init__(rc)
        # TODO: templates for other universities?
        self.template = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates", "reimb.xlsx")
        self.cmds = ["excel"]

    def construct_global_ctx(self):
        """Constructs the global context."""
        super().construct_global_ctx()
        gtx = self.gtx
        rc = self.rc
        # openpyxl is soooo slow, so only allow it to be run when a person
        # (or people list) is specified
        if not rc.people:
            raise ValueError(
                "Missing person for the reimbursement.  Please "
                "rerun specifying --people and a person or list "
                "of people"
            )
        gtx["month_and_year"] = month_and_year
        gtx["people"] = sorted(
            all_docs_from_collection(rc.client, "people"),
            key=position_key,
            reverse=True,
        )
        for n in ["expenses", "grants"]:
            gtx[n] = list(all_docs_from_collection(rc.client, n))
        gtx["all_docs_from_collection"] = all_docs_from_collection

    def excel(self):
        gtx = self.gtx
        rc = self.rc
        if isinstance(rc.people, str):
            rc.people = [rc.people]
        for ex in gtx["expenses"]:
            payee = fuzzy_retrieval(gtx["people"], ["name", "aka", "_id"], ex["payee"])
            chosen_ones = [fuzzy_retrieval(gtx["people"], ["name", "aka", "_id"], one) for one in rc.people]
            if ex["payee"] != "direct_billed":
                for chosen_one in chosen_ones:
                    if not payee:
                        print(f"WARNING: payee {ex['payee']} not found in " f"people coll")
                    elif payee.get("name") != chosen_one.get("name"):
                        continue
                    # open the template
                    if isinstance(ex["grants"], str):
                        ex["grants"] = [ex["grants"]]
                        grant_fractions = [1.0]
                    else:
                        grant_fractions = [float(percent) / 100.0 for percent in ex["grant_percentages"]]

                    wb = openpyxl.load_workbook(self.template)
                    ws = wb["T&B"]

                    grants = []
                    for grant_label in ex["grants"]:
                        grant = fuzzy_retrieval(gtx["grants"], ["alias", "name", "_id"], grant_label)
                        if not grant:
                            raise ValueError(f"no grant found with label {grant_label}")
                        else:
                            grants.append(grant)
                    ha = payee["home_address"]
                    ws["B17"] = payee["name"]
                    ws["B20"] = ha["street"]
                    ws["B23"] = ha["city"]
                    ws["G23"] = ha["state"]
                    ws["L23"] = ha["zip"]
                    ws["B36"] = ex["overall_purpose"]
                    j = 42
                    total_amount = 0
                    item_ws = wb["T&B"]
                    purpose_column = 4
                    ue_column = 13
                    se_column = 16
                    dates = []
                    for i, item in enumerate(ex["itemized_expenses"]):
                        r = j + i
                        expdates = get_dates(item)
                        if r > 49:
                            item_ws = wb["Extra_Page"]
                            j = 0
                            r = j + i
                            purpose_column = 5
                            ue_column = 12
                            se_column = 14
                        dates.append(expdates.get("date"))
                        item_ws.cell(row=r, column=2, value=i)
                        item_ws.cell(row=r, column=3, value=expdates.get("date").strftime("%x"))
                        item_ws.cell(row=r, column=purpose_column, value=item["purpose"])
                        item_ws.cell(
                            row=r,
                            column=ue_column,
                            value=item.get("unsegregated_expense", 0),
                        )
                        try:
                            total_amount += item.get("unsegregated_expense", 0)
                        except TypeError:
                            if item.get("unsegregated_expense", 0) == "tbd":
                                print("WARNING: unsegregated expense in {} is " "tbd".format(ex["_id"]))
                                item["unsegregated_expense"] = 0
                            else:
                                raise TypeError("unsegregated expense in {} is not " "a number".format(ex["_id"]))

                        item_ws.cell(row=r, column=se_column, value=item.get("segregated_expense", 0))

                    i = 0
                    if abs(sum([fraction * total_amount for fraction in grant_fractions]) - total_amount) >= 0.01:
                        raise RuntimeError("grant percentages do not sum to 100")
                    for grant, fraction in zip(grants, grant_fractions):
                        nr = grant.get("account", "")
                        row = 55 + i
                        location = "C{}".format(row)
                        location2 = "K{}".format(row)
                        ws[location] = nr
                        ws[location2] = total_amount * float(fraction)
                        i += 1

                    if ex.get("expense_type", "business") == "business":
                        spots = ("G10", "L11", "O11")
                    else:
                        spots = ("G7", "L8", "O8")

                    ws[spots[0]] = "X"
                    ws[spots[1]] = min(dates).strftime("%x")
                    ws[spots[2]] = max(dates).strftime("%x")

                    wb.save(os.path.join(self.bldir, ex["_id"] + ".xlsx"))
