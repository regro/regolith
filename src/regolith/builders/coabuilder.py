"""Builder for Recent Collaborators.

For the specified person it returns the name and institution of all
graduate student and post-doc advisors, all graduate student advisees,
all post-doc advisees in the past 60 months and all coauthors in the
past 48 months.
"""

import datetime as dt
import os
import sys
from copy import copy

import openpyxl
from dateutil.relativedelta import relativedelta
from nameparser import HumanName

from regolith.builders.basebuilder import BuilderBase
from regolith.dates import get_dates, is_after, month_to_int
from regolith.sorters import position_key
from regolith.tools import all_docs_from_collection, filter_publications, fuzzy_retrieval

NUM_COAUTHOR_MONTHS = 48
NUM_POSTDOC_MONTHS = None


def get_advisors_name_inst(advisee, rc):
    """Get the advisee's advisor.

    Yield (last name, first name, institution name).
    """

    phd_advisors = [
        {
            "name": i.get("advisor", "missing name"),
            "type": "advisor",
            "advis_type": "phd",
            "interaction_date": get_dates(i).get("end_date", get_dates(i).get("begin_date")),
        }
        for i in advisee.get("education", [])
        if "phd" in i.get("degree", "").lower() or "dphil" in i.get("degree", "").lower()
    ]
    pdoc_advisors = [
        {
            "name": i.get("advisor", "missing name"),
            "type": "advisor",
            "advis_type": "postdoc",
            "interaction_date": get_dates(i).get("end_date", get_dates(i).get("begin_date")),
        }
        for i in advisee.get("employment", [])
        if i.get("status") == "postdoc"
    ]
    advisors = phd_advisors + pdoc_advisors
    return retrieve_names_and_insts(rc, advisors)


def get_advisees_name_inst(coll, advisor, rc):
    """Get advisor's advisees.

    Yield (last name, first name, institutions)
    """
    advisor_names = advisor.get("aka", []) + [advisor.get("name"), advisor.get("_id")]
    advisees = []
    for person in coll:
        my_eme = person.get("employment", []) + person.get("education", [])
        relevant_emes = [i for i in my_eme if i.get("advisor", "") in advisor_names]
        phd_advisees = [
            {
                "name": person.get("name", "missing name"),
                "type": "advisee",
                "interaction_date": get_dates(i).get("end_date", get_dates(i).get("date", dt.date.today())),
                "advis_type": "phd",
            }
            for i in relevant_emes
            if "phd" in i.get("degree", "").lower() or "dphil" in i.get("degree", "").lower()
        ]
        pdoc_advisees = [
            {
                "name": person.get("name", "missing name"),
                "type": "advisee",
                "advis_type": "postdoc",
                "interaction_date": get_dates(i).get("end_date", get_dates(i).get("date", dt.date.today())),
            }
            for i in relevant_emes
            if i.get("status") == "postdoc"
        ]
        if rc.postdoc_since_date:
            pdoc_advisees = [i for i in pdoc_advisees if rc.postdoc_since_date < i.get("interaction_date")]
        advisees.extend(phd_advisees)
        advisees.extend(pdoc_advisees)
    return retrieve_names_and_insts(rc, advisees)


def filter_since_date(pubs, rc):
    """Filter the publications after the since_date."""
    for pub in pubs:
        if isinstance(pub.get("year"), str):
            pub["year"] = int(pub.get("year"))
        pub["day"] = int(pub.get("day", 28))
        pub["month"] = pub.get("month", "dec")
        if pub.get("month").casefold().strip() == "tbd" or pub.get("month").strip() == "":
            print("WARNING: {} is missing month".format(pub["_id"]))
            pub["month"] = "dec"
        if is_after(pub, rc.pub_since_date):
            yield pub


def get_since_dates(rc):
    """Get the since_date from command."""
    if isinstance(rc.date, str):
        rc.pub_since_date = dt.datetime.strptime(rc, "%Y-%m-%d").date() - relativedelta(months=NUM_COAUTHOR_MONTHS)
    else:
        rc.pub_since_date = rc.date - relativedelta(months=NUM_COAUTHOR_MONTHS)
    if NUM_POSTDOC_MONTHS:
        if isinstance(rc, str):
            rc.postdoc_since_date = dt.datetime.strptime(rc.date, "%Y-%m-%d").date() - relativedelta(
                months=NUM_POSTDOC_MONTHS
            )
        else:
            rc.pub_since_date = rc.date - relativedelta(months=NUM_COAUTHOR_MONTHS)
    else:
        rc.postdoc_since_date = None
    return


def get_coauthors_from_pubs(rc, pubs, not_person):
    """Get co-authors' names from the publication.

    Not include the person itself.
    """
    not_person_akas = [not_person["_id"], not_person["name"]] + not_person["aka"]
    my_collabs = list()
    for pub in pubs:
        pub_date = dt.date(int(pub.get("year")), month_to_int(pub.get("month")), 1)
        my_collabs.extend(
            [
                {"name": collabs, "interaction_date": pub_date, "type": "co-author"}
                for collabs in (names for names in pub.get("author", []))
            ]
        )
    my_collabs.sort(key=lambda x: x["interaction_date"], reverse=True)
    coauthors = retrieve_names_and_insts(rc, my_collabs, not_person_akas)
    return coauthors


def retrieve_names_and_insts(rc, collabs, not_person_akas=[]):
    collab_buffer, my_collab_set = [], []
    for collab in collabs:
        person = fuzzy_retrieval(
            all_docs_from_collection(rc.client, "people"),
            ["name", "aka", "_id"],
            collab["name"],
            case_sensitive=False,
        )
        if not person:
            person = fuzzy_retrieval(
                all_docs_from_collection(rc.client, "contacts"),
                ["name", "aka", "_id"],
                collab["name"],
                case_sensitive=False,
            )
            if not person:
                if collab["name"] == "missing name":
                    print(
                        f"WARNING: a {collab.get('advis_type')} appointment "
                        f"was found for the target {collab.get('type')} but "
                        f"no name was specified. Please add an 'advisor' field "
                        f"for that education/employment entry in the database."
                    )
                else:
                    print(f"WARNING: {collab['name']} not found in contacts or people.")
                person = {"_id": collab["name"], "name": collab["name"], "type": collab.get("type")}
        if person.get("name", ""):
            collab["name"] = HumanName(person.get("name", ""))
        else:
            print("missing_person", person)
        collab["_id"] = person.get("_id")
        pinst = get_recent_org(person)
        inst = fuzzy_retrieval(
            all_docs_from_collection(rc.client, "institutions"),
            ["name", "aka", "_id"],
            pinst,
            case_sensitive=False,
        )
        if inst:
            collab["institution"] = inst["name"]
        else:
            collab["institution"] = pinst
            print(f"WARNING: {pinst} for {person.get('_id')} missing from institutions")
        if collab["_id"] not in collab_buffer and collab["name"] not in not_person_akas:
            my_collab_set.append(collab)
            collab_buffer.append(collab["_id"])
    return my_collab_set


def get_recent_org(person_info):
    """Get the person's most recent organization."""
    if "institution" in person_info:
        organization = person_info.get("institution")
    elif "employment" in person_info:
        employment = person_info.get("employment", []) + person_info.get("education", [])
        if len(employment) == 0:
            return ""
        # sort by end_year
        employment = sorted(employment, key=lambda d: d.get("end_year", float("inf")), reverse=True)
        organization = employment[0].get("organization", employment[0].get("institution"))
    else:
        organization = ""
    return organization


def query_people_and_institutions(rc, names):
    """Get the people and institutions names."""
    people, institutions, latest_active = [], [], []
    for person_name in names:
        person_found = fuzzy_retrieval(
            all_docs_from_collection(rc.client, "people"),
            ["name", "aka", "_id"],
            person_name[0],
            case_sensitive=False,
        )
        if not person_found:
            person_found = fuzzy_retrieval(
                all_docs_from_collection(rc.client, "contacts"),
                ["name", "aka", "_id"],
                person_name[0],
                case_sensitive=False,
            )
            if not person_found:
                print(
                    "WARNING: {} not found in contacts or people. Check aka".format(person_name[0]).encode("utf-8")
                )
            else:
                people.append(person_found["name"])
                inst = fuzzy_retrieval(
                    all_docs_from_collection(rc.client, "institutions"),
                    ["name", "aka", "_id"],
                    person_found["institution"],
                    case_sensitive=False,
                )
                if inst:
                    institutions.append(inst["name"])
                else:
                    institutions.append(person_found.get("institution", "missing"))
                    print("WARNING: {} missing from institutions".format(person_found["institution"]))
        else:
            people.append(person_found["name"])
            pinst = get_recent_org(person_found)
            inst = fuzzy_retrieval(
                all_docs_from_collection(rc.client, "institutions"),
                ["name", "aka", "_id"],
                pinst,
                case_sensitive=False,
            )
            if inst:
                institutions.append(inst["name"])
            else:
                institutions.append(pinst)
                print("WARNING: {} missing from institutions".format(pinst))
        latest_active.append(person_name[1])
    return people, institutions, latest_active


def get_inst_name(person, rc):
    """Get the name of institution of the person's latest employment."""
    if "employment" in person:
        org = get_recent_org(person)
        person_inst_abbr = org
    elif "institution" in person:
        person_inst_abbr = person.get("institution")
    else:
        person_inst_abbr = ""
    person_inst = fuzzy_retrieval(
        all_docs_from_collection(rc.client, "institutions"),
        ["name", "aka", "_id"],
        person_inst_abbr,
        case_sensitive=False,
    )
    if person_inst is not None:
        person_inst_name = person_inst.get("name")
    else:
        person_inst_name = person_inst_abbr
        print(f"WARNING: {person_inst_abbr} is not found in institutions.")
    return person_inst_name


def get_person_pubs(coll, person):
    """Get the publications from one person."""
    my_names = frozenset(person.get("aka", []) + [person["name"]])
    pubs = filter_publications(coll, my_names, reverse=True, bold=False)
    return pubs


def make_person_4tups(person, rc):
    if "name" not in person:
        print("Warning")
    name = HumanName(person["name"])
    inst = get_inst_name(person, rc)
    first_names = " ".join([name.first, name.middle])
    return [(name.last, first_names, inst, "pi")]


def format_last_first_institution_names(rc, ppl_names, excluded_inst_name=None):
    """Get the last name, first name and institution name."""
    ppl = []
    for ppl_tup in ppl_names:
        inst = fuzzy_retrieval(
            all_docs_from_collection(rc.client, "institutions"),
            ["aka", "name", "_id"],
            ppl_tup[1],
            case_sensitive=False,
        )
        if inst:
            inst_name = inst.get("name", "")
        else:
            inst_name = ppl_tup[1]
        # remove all people who are in the institution of the person
        if inst_name != excluded_inst_name:
            name = HumanName(ppl_tup[0])
            yield name.last, " ".join([name.first, name.middle]), ppl_tup[1], " ", ppl_tup[2]
    return ppl


def format_to_nsf(tups, type_str):
    """Format the 3 tups to 2 tups.

    ('type_str', 'last, first', 'inst', ...).
    """
    return [(type_str, "{}, {}".format(tup[0], tup[1])) + tup[2:] for tup in tups]


def apply_cell_style(*cells, style):
    """Apply the format to cells."""
    for cell in cells:
        cell.font = style["font"]
        cell.border = style["border"]
        cell.fill = style["fill"]
        cell.alignment = style["alignment"]
    return


def copy_cell_style(style_ref_cell):
    """Copy the cell format to a dictionary."""
    template_cell_style = {
        "font": copy(style_ref_cell.font),
        "border": copy(style_ref_cell.border),
        "fill": copy(style_ref_cell.fill),
        "alignment": copy(style_ref_cell.alignment),
    }
    return template_cell_style


def is_merged(cell):
    """Whether the cell is merged."""
    return True if type(cell).__name__ == "MergedCell" else False


def find_merged_cell(cells):
    """Find the index of all merged cells."""
    i, j = 0, 0
    lst = []
    while i < len(cells):
        if is_merged(cells[i]):
            while j < len(cells):
                if is_merged(cells[j]):
                    j += 1
                else:
                    i = j
                    break
            lst.append((i, j - 1))
        else:
            i += 1
    return lst


def unmerge(ws, cells):
    """Unmerge the cells."""
    lst = find_merged_cell(cells)
    for start, end in lst:
        ws.unmerge_cells("{}:{}").format(cells[start].coordinate, cells[end].coordinate)
    return


def get_person(person_id, rc):
    """Get the person's name."""
    person_found = fuzzy_retrieval(
        all_docs_from_collection(rc.client, "people"), ["name", "aka", "_id"], person_id, case_sensitive=False
    )
    if person_found:
        return person_found
    person_found = fuzzy_retrieval(
        all_docs_from_collection(rc.client, "contacts"), ["name", "aka", "_id"], person_id, case_sensitive=False
    )
    if not person_found:
        print("WARNING: {} missing from people and contacts. Check aka.".format(person_id))
        person_found = {"name": person_id}
    return person_found


def find_coeditors(person, rc):
    """Get the coeditors info of the person.

    Return (last, first, inst, journal).
    """
    emps = person.get("employment")
    if emps is None:
        return set()

    def coeditor_id_journals(_emps):
        for emp in _emps:
            if emp.get("position") == "editor":
                _journal = emp.get("department", "")
                coeditor_ids = emp.get("coworkers", [])
                for _coeditor_id in coeditor_ids:
                    yield _coeditor_id, _journal

    coeditor_inst_journals = set()
    for coeditor_id, journal in coeditor_id_journals(emps):
        coeditor = get_person(coeditor_id, rc)
        coeditor_name = HumanName(coeditor.get("name"), "")
        inst_name = get_inst_name(coeditor, rc)
        coeditor_inst_journals.add((coeditor_name.last, coeditor_name.first, inst_name, journal, "co-editor"))
    return coeditor_inst_journals


class RecentCollaboratorsBuilder(BuilderBase):
    """Build recent collaborators from database entries."""

    btype = "recent-collabs"
    needed_colls = ["citations", "people", "contacts", "institutions"]

    def __init__(self, rc):
        """Initiate the class instance."""
        super().__init__(rc)
        self.template = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "templates", "coa_template_nsf.xlsx"
        )
        self.template2 = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "templates", "coa_template_doe.xlsx"
        )
        self.cmds = ["excel"]
        try:
            rc.verbose
        except AttributeError:
            rc.verbose = False

    def construct_global_ctx(self):
        """Construct the global ctx including database and methods."""
        super().construct_global_ctx()
        gtx = self.gtx
        rc = self.rc
        gtx["people"] = sorted(
            all_docs_from_collection(rc.client, "people"),
            key=position_key,
            reverse=True,
        )
        gtx["contacts"] = sorted(
            all_docs_from_collection(rc.client, "contacts"),
            key=position_key,
            reverse=True,
        )
        gtx["institutions"] = all_docs_from_collection(rc.client, "institutions")
        gtx["citations"] = all_docs_from_collection(rc.client, "citations")
        gtx["all_docs_from_collection"] = all_docs_from_collection

    def query_ppl(self, target):
        """Query the data base for the target's collaborators'
        information."""
        rc = self.rc
        gtx = self.gtx
        person = fuzzy_retrieval(
            all_docs_from_collection(rc.client, "people"), ["aka", "name", "_id"], target, case_sensitive=False
        )
        if not person:
            raise RuntimeError("Person {} not found in people.".format(target).encode("utf-8"))

        pubs = get_person_pubs(gtx["citations"], person)
        pubs = filter_since_date(pubs, rc)
        try:
            if rc.verbose:
                for pub in pubs:
                    print(f"{pub.get('title')}, ({pub.get('year')})")
        except AttributeError:
            pass
        my_collabs = get_coauthors_from_pubs(rc, pubs, person)
        advisors = get_advisors_name_inst(person, rc)
        advisees = get_advisees_name_inst(all_docs_from_collection(rc.client, "people"), person, rc)
        collabs = []
        adviseors = advisors + advisees
        for collab in my_collabs:
            col_bool = True
            for advis in adviseors:
                if (
                    collab.get("name").last == advis.get("name").last
                    and collab.get("name").first == advis.get("name").first
                ):
                    col_bool = False
                    if advis.get("interaction_date"):
                        if collab.get("interaction_date", dt.date.today()) > advis.get(
                            "interaction_date", dt.date.today()
                        ):
                            advis.update({"interaction_date": collab.get("interaction_date")})
                    try:
                        if collab.get("interaction_date") > advis.get("interaction_date"):
                            advis.update({"interaction_date": collab.get("interaction_date")})
                    except TypeError:
                        print(f"ERROR: incorrect dates for an education/employment in {collab.get('name')}")
                        print(
                            f"collab date: {collab.get('interaction_date')}, "
                            f"advisee date: {advis.get('interaction_date')}"
                        )
                        raise
            if col_bool:
                collabs.append(collab)
        collabs.extend(advisees)
        collabs.extend(advisors)
        collabs.sort(key=lambda d: d["name"].last)
        if rc.verbose:
            output = [
                f"{my_collab.get('name').last}, "
                f"{my_collab.get('name').first}, "
                f"{my_collab.get('institution')}, "
                f"{my_collab.get('interaction_date')}, "
                f"{my_collab.get('advis_type', '')}, "
                f"{my_collab.get('type')}\n"
                for my_collab in collabs
            ]
            print(*output)
        person["name"] = HumanName(person.get("name"))
        results = {"person_info": person, "collabs": collabs}
        return results

    @staticmethod
    def fill_in_tab(ws, ppl, start_row, template_cell_style=None, cols="ABCDE"):
        """Add the information in person, institution pairs into the
        table 4 in nsf table."""
        nsf_mappings = {
            "co-author": "A:",
            "collaborator": "C:",
            "phd_advisor": "G:",
            "phd_advisee": "T:",
            "co-editor": "E:",
        }
        more_rows = len(ppl) - 1
        if more_rows > 0:
            ws.insert_rows(start_row, amount=more_rows)
        for row, person in enumerate(ppl, start=start_row):
            cells = [ws[f"{col}{row}"] for col in cols]
            if template_cell_style is not None:
                apply_cell_style(*cells, style=template_cell_style)
                ws[f"A{row}"].value = nsf_mappings.get(person.get("type"))
                ws[f"B{row}"].value = f"{person.get('name').last}, " f"{person.get('name').first}"
                ws[f"C{row}"].value = person.get("institution")
                if isinstance(person.get("interaction_date"), dt.date):
                    ws[f"E{row}"].value = person.get("interaction_date", dt.date.today()).strftime("%m/%d/%Y")
        return

    def render_template1(self, person_info, collabs, **kwargs):
        """Render the nsf template."""
        template = self.template
        wb = openpyxl.load_workbook(template)
        ws = wb.worksheets[0]
        nsf_collabs = copy(collabs)
        advis, coauths, coeditors = [], [], []
        pi_row_number = 17
        ws[f"B{pi_row_number}"].value = f"{person_info.get('name').last}, " f"{person_info.get('name').first}"
        ws[f"C{pi_row_number}"].value = person_info.get("institution")
        ws[f"D{pi_row_number}"].value = person_info.get("interaction_date", dt.date.today()).strftime("%m/%d/%Y")

        for collab in nsf_collabs:
            if collab.get("type") == "advisor" and collab.get("advis_type") == "phd":
                collab.update({"type": "phd_advisor"})
                advis.append(collab)
            if collab.get("type") == "advisee" and collab.get("advis_type") == "phd":
                collab.update({"type": "phd_advisee"})
                advis.append(collab)
            if collab.get("type") == "co-author":
                coauths.append(collab)
            if collab.get("type") == "co-editor":
                coeditors.append(collab)
        style = copy_cell_style(ws["A17"])

        if coauths:
            self.fill_in_tab(ws, coauths, start_row=52, template_cell_style=style)
        if advis:
            self.fill_in_tab(ws, advis, start_row=38, template_cell_style=style)
        # if person_info:
        #     self.fill_in_tab(
        #         ws, [person_info], start_row=17, template_cell_style=style
        #     )
        if coeditors:
            self.fill_in_tab(ws, coeditors, start_row=120, template_cell_style=style)
        wb.save(os.path.join(self.bldir, "{}_nsf.xlsx".format(person_info["_id"])))
        return locals()

    def render_template2(self, person_info, collabs, **kwargs):
        """Render the doe template."""

        template2 = self.template2
        num_rows = len(collabs)
        wb = openpyxl.load_workbook(template2)
        ws = wb.worksheets[0]
        for row in range(num_rows):
            ws["A{}".format(row + 8)].value = person_info["name"].last
            ws["B{}".format(row + 8)].value = person_info["name"].first
            ws["C{}".format(row + 8)].value = collabs[row]["name"].last
            ws["D{}".format(row + 8)].value = collabs[row]["name"].first
            ws["F{}".format((row + 8))].value = collabs[row]["institution"]
            ws["G{}".format((row + 8))].value = collabs[row]["type"]
            if isinstance(collabs[row]["interaction_date"], dt.date):
                ws["H{}".format((row + 8))].value = collabs[row]["interaction_date"].year
        ws["C3"].value = person_info["name"].full_name
        ws["C4"].value = person_info.get("email", "")
        wb.save(os.path.join(self.bldir, "{}_doe.xlsx".format(person_info["_id"])))
        return locals()

    def excel(self):
        """Query data base and build nsf and doe excels."""
        rc = self.rc
        if not rc.people:
            sys.exit("please rerun specifying --people PERSON")
        if isinstance(rc.people, str):
            rc.people = [rc.people]
        if not hasattr(rc, "date"):
            rc.date = dt.date.today()
        get_since_dates(rc)
        rc.post_doc_window_months = NUM_POSTDOC_MONTHS
        print(f"filtering coauthors for papers since {rc.pub_since_date}")
        for target in rc.people:
            query_results = self.query_ppl(target)
            self.render_template1(**query_results)
            self.render_template2(**query_results)
